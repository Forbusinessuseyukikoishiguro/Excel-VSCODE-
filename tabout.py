from openpyxl import load_workbook

# Excelファイルのパスを指定
excel_path = r"C:\Users\yukik\Downloads\モジュール4の英訳ディープAIこーせら .xlsx"

# Excelを読み込み
wb = load_workbook(excel_path)

# 作成するシートの数
num_sheets = 45

# 連番で「モジュール1」「モジュール2」…というシートを追加
for i in range(1, num_sheets + 1):
    sheet_name = f"モジュール{i}"
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name)

# 保存
wb.save(excel_path)
print("✅ モジュール1〜モジュール45 のタブシートを指定したExcelファイルにて作成したにゃー(=^・^=)")



#tabout.py
